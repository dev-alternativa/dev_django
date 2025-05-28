import os
import re
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from django.core.management.base import BaseCommand
from common.models import CustomerSupplier, CNPJFaturamento, Seller
from logistic.models import LeadTime, Freight
from products.models import Price, Product


class Command(BaseCommand):
    '''
    Importa preços de clientes da planilha de Excel "template_tabela_precos.xlsx"

    Este comando lê uma planilha Excel contendo preços de clientes e importa
    os dados para o banco de dados. Lida com a criação de objetos associados, como
    Product, CustomerSupplier, LeadTime, CNPJFaturamento e Seller.

    Métodos:
        add_arguments(self, parser): Define argumentos de linha de comando ao comando.
        handle(self, *args, **options): Manipula a lógica principal do comando.
        _import_from_excel(self, file_path, batch_size): Importa dados da planilha Excel.
        _create_object_from_row(self, row): Cria um objeto Price a partir de uma linha de dados.
        _normalize_string(s): Normaliza uma string removendo espaços extras e substituindo vírgulas por pontos.
        _round_decimal(value, places): Arredonda um valor decimal para o número especificado de casas decimais.
    '''
    help = 'Importa preços de clientes da planilha excel "template_tabela_precos.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Caminho do arquivo excel')
        parser.add_argument('--batch-size', type=int, default=1000, help='Número de registros para inserir por vez')


    def handle(self, *args, **options):
        file_path = options['file_path']
        batch_size = options['batch_size']

        self.stdout.write(f'Importando dados de {file_path}...')

        # Determina se o arquivo tem extensão .xlsx válida
        _, file_extension = os.path.splitext(file_path)

        if file_extension.lower() in ['.xlsx', '.xls']:
            self._import_from_excel(file_path, batch_size)
        else:
            self.stdout.write(self.style.ERROR(f'Formato de arquivo {file_extension} é inválido!'))

    def _import_from_excel(self, file_path, batch_size):
        # Countador de registros
        errors = 0
        count = 0
        batch = []

        # Carrega a 1 planilha por padrão
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook['Tab_Import']

        # Obtém os cabeçalhos da planilha
        headers = [cell.value for cell in next(sheet.iter_rows())]

        # Itera pelas linhas de dados (a partir da 2ª linha)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
            try:
                # Converte a linha em um dicionário
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                if row_data['Tabela'] != 'GRAV':
                    obj = self._create_object_from_row(row_data)
                    if obj:
                        batch.append(obj)
                        count += 1
                    else:
                        errors += 1
                    # if len(batch) >= batch_size:
                    #     Price.objects.bulk_create(batch)
                    #     self.stdout.write(f'Importados {count} registros...')
                    #     batch = []
            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f'Erro ao importar linha {row_idx}: {str(e)}'))

        print(batch[0].__dict__)

        # if batch:
            # Price.objects.bulk_create(batch)

        workbook.close()
        self.stdout.write(self.style.SUCCESS(f'Importação concluída! {count} registros inseridos, {errors} erros.'))

    @staticmethod
    def _normalize_string(s):
        s = s.replace('  ', '').replace(',', '.').strip()
        return s

    @staticmethod
    def _round_decimal(value, places=2):
        if value is None:
            return None

        str_value = str(value)
        decimal_value = Decimal(str_value)

        format = Decimal('0.01')
        if places != 2:
            format = Decimal('0.' + '0' * places)

        return decimal_value.quantize(format, rounding=ROUND_HALF_UP)


    def _create_object_from_row(self, row):
        '''
        Cria um objeto Price a partir de uma linha de dados.
        Lida com as chaves estrangeiras buscando os objetos correspondentes.
        '''
        try:
            # PRODUTO
            sheet_product = self._normalize_string(row.get('PRODUTO'))
            product = Product.objects.get(nome_produto=sheet_product)


            # CLIENTE
            cnpj = row.get('CNPJ CLIENTE').replace('.', '').replace('/', '').replace('-', '').strip()
            customer = CustomerSupplier.objects.get(cnpj=cnpj)

            # PRAZO
            lead_time = None
            if prazo := row.get('PRAZO'):
                try:
                    # print(prazo)
                    normalized_prazo = self._normalize_string(prazo)
                    # print(normalized_prazo)
                    lead_time = LeadTime.objects.get(descricao=normalized_prazo)
                except LeadTime.DoesNotExist:
                    pass

            # CNPJ_FATURAMENTO
            cnpj = row.get('CNPJ_FATURAMENTO')
            cnpj_fat_obj = CNPJFaturamento.objects.get(sigla=cnpj)

            # VENDEDOR
            seller = row.get('VENDEDOR')
            seller_obj = Seller.objects.get(nome=seller)

            # VALOR
            value = Decimal(str(row.get('VALOR', 0)))

            # FRETE
            shipping_value = None
            if frete_valor := row.get('VALOR FRETE'):
                shipping_value = Decimal(str(frete_valor))

            # DOLAR
            is_dolar = str(row.get('DOLAR', '')) in ['1', 'sim', 's', 'true']

            # TIPO FRETE
            tipo_frete = None
            if freight_type := row.get('Tipo de frete'):
                try:
                    tipo_frete = Freight.objects.get(tipo_frete__icontains=freight_type).tipo_frete
                except Freight.DoesNotExist:
                    pass

            # Mapeamento dos campos/colunas da tabela
            return Price(
                produto=product,
                cliente=customer,
                valor=value,
                is_dolar=is_dolar,
                prazo=lead_time,
                cnpj_faturamento=cnpj_fat_obj,
                condicao=row.get('CONDICAO'),
                vendedor=seller_obj,
                frete=shipping_value,
                tipo_frete=tipo_frete,
                obs=row.get('OBS'),
            )

        except Product.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'Produto não encontrado: {row.get("PRODUTO", '')}'))

        except CustomerSupplier.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'Cliente não encontrado: {row.get('CLIENTE NOME FANTASIA')}'))
        except CNPJFaturamento.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'CNPJ para Faturamento não encontrado: {row.get('CNPJ FATURAMENTO')}'))
        except Seller.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'Vendedor não encontrado: {row.get("VENDEDOR")}'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Erro ao processar linha: {str(e)}'))

        return None # Ignora a linha com erro